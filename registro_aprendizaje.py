import tkinter as tk
from tkinter import messagebox, font
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import msoffcrypto
import io
import win32com.client

# Configuración local
ruta_excel = r"C:\Users\Ricrado Perez\Desktop\Aprendizaje\registro.xlsx"
contraseña_excel = "d8104dd8"  # Reemplaza por tu contraseña real

def centrar_ventana(ventana):
    ventana.update_idletasks()
    ancho = ventana.winfo_width()
    alto = ventana.winfo_height()
    x = (ventana.winfo_screenwidth() // 2) - (ancho // 2)
    y = (ventana.winfo_screenheight() // 2) - (alto // 2)
    ventana.geometry(f'+{x}+{y}')

def cargar_excel_con_contraseña(ruta, contraseña):
    with open(ruta, 'rb') as f:
        archivo_encriptado = msoffcrypto.OfficeFile(f)
        archivo_encriptado.load_key(password=contraseña)
        archivo_desencriptado = io.BytesIO()
        archivo_encriptado.decrypt(archivo_desencriptado)
        archivo_desencriptado.seek(0)
        return pd.read_excel(archivo_desencriptado)

def guardar_excel_sin_proteccion(df, ruta):
    temp_path = ruta.replace(".xlsx", "_temp.xlsx")
    df.to_excel(temp_path, index=False)
    os.replace(temp_path, ruta)

def proteger_excel_contraseña(ruta, contraseña):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(ruta, False, False, None, contraseña)
    wb.SaveAs(ruta, None, contraseña)  # Reaplica la contraseña
    wb.Close()
    excel.Quit()

def inicializar_excel():
    if not os.path.exists(ruta_excel):
        df = pd.DataFrame(columns=[
            "Fecha", "Aprendizaje", "Logro", "Mejorar",
            "Estado emocional", "Valoración emocional", "Hábitos"
        ])
        df.to_excel(ruta_excel, index=False)

def dar_formato_excel():
    wb = load_workbook(ruta_excel)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max(20, min(max_length + 2, 100))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(ruta_excel)

def guardar_respuesta():
    aprendizaje = entry_aprendizaje.get("1.0", tk.END).strip()
    logro = entry_logro.get("1.0", tk.END).strip()
    mejorar = entry_mejorar.get("1.0", tk.END).strip()
    estado = var_estado.get()
    habitos = entry_habitos.get("1.0", tk.END).strip()

    if not (aprendizaje and logro and mejorar and estado):
        messagebox.showwarning("Campos incompletos", "Por favor, llena todos los campos.")
        return

    emociones = {
        1: "😞 1 - Muy mal", 2: "😕 2 - Mal", 3: "😐 3 - Regular", 4: "🙂 4 - Algo bien",
        5: "😊 5 - Bien", 6: "😃 6 - Muy bien", 7: "😁 7 - Excelente", 
        8: "🤩 8 - Fantástico", 9: "🥰 9 - Increíble", 10: "😍 10 - Perfecto"
    }

    nueva_fila = {
        "Fecha": datetime.now().strftime("%Y-%m-%d"),
        "Aprendizaje": aprendizaje,
        "Logro": logro,
        "Mejorar": mejorar,
        "Estado emocional": estado,
        "Valoración emocional": emociones[estado],
        "Hábitos": habitos
    }

    df = cargar_excel_con_contraseña(ruta_excel, contraseña_excel)
    df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
    guardar_excel_sin_proteccion(df, ruta_excel)
    dar_formato_excel()
    proteger_excel_contraseña(ruta_excel, contraseña_excel)
    messagebox.showinfo("Guardado", "Registro exitoso!")
    root.destroy()

def on_tab(event):
    widget = event.widget
    tab_order = [
        entry_aprendizaje,
        entry_logro,
        entry_mejorar,
        *radio_buttons,
        entry_habitos,
        btn_guardar
    ]
    try:
        index = tab_order.index(widget)
        next_widget = tab_order[(index + 1) % len(tab_order)]
        if next_widget in radio_buttons:
            var_estado.set(next_widget["value"])
        next_widget.focus_set()
        if isinstance(next_widget, tk.Text):
            next_widget.mark_set("insert", "1.0")
    except ValueError:
        pass
    return "break"

# --- INTERFAZ GRÁFICA ---
root = tk.Tk()
root.title("Mi Diario de Aprendizaje")
root.geometry("650x850")
root.configure(bg='#1e1e1e')

# Fuentes
try:
    titulo_font = font.Font(family="Segoe UI", size=12, weight="bold")
    label_font = font.Font(family="Segoe UI Light", size=11)
    entry_font = font.Font(family="Segoe UI", size=10)
    button_font = font.Font(family="Segoe UI Semibold", size=10)
    emoji_font = font.Font(family="Segoe UI Emoji", size=10)
except:
    titulo_font = font.Font(size=12, weight="bold")
    label_font = font.Font(size=11)
    entry_font = font.Font(size=10)
    button_font = font.Font(size=10, weight="bold")
    emoji_font = font.Font(size=10)

bg_color = '#1e1e1e'
label_color = '#e0e0e0'
entry_bg = '#2e2e2e'
entry_fg = '#ffffff'
button_bg = '#3a3a3a'
button_fg = '#ffffff'
button_active = '#4a4a4a'
radio_selected = '#64b5f6'
radio_bg = '#2e2e2e'

main_frame = tk.Frame(root, bg=bg_color)
main_frame.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

tk.Label(main_frame, text="Resumen Diario", bg=bg_color, fg='#64b5f6', font=titulo_font).pack(pady=(0, 20))

def crear_campo(marco, texto, es_texto=True, height=3):
    tk.Label(marco, text=texto, bg=bg_color, fg=label_color, font=label_font, anchor='w').pack(fill=tk.X, pady=(10, 0))
    if es_texto:
        entry = tk.Text(marco, height=height, width=60, bg=entry_bg, fg=entry_fg,
                        insertbackground='white', font=entry_font, wrap=tk.WORD, padx=5, pady=5)
        entry.pack(pady=5)
    else:
        entry = tk.Entry(marco, width=10, bg=entry_bg, fg=entry_fg, insertbackground='white', font=entry_font)
        entry.pack(pady=5)
    return entry

entry_aprendizaje = crear_campo(main_frame, "1. Lo más valioso que aprendí hoy:")
entry_logro = crear_campo(main_frame, "2. Mi mayor logro del día:")
entry_mejorar = crear_campo(main_frame, "3. Algo que puedo mejorar:")

tk.Label(main_frame, text="4. Estado emocional (1-10):", bg=bg_color, fg=label_color, font=label_font).pack(pady=(10, 0), anchor='w')
emociones_frame = tk.Frame(main_frame, bg=bg_color)
emociones_frame.pack()

var_estado = tk.IntVar()
emociones = {1:"😞", 2:"😕", 3:"😐", 4:"🙂", 5:"😊", 6:"😃", 7:"😁", 8:"🤩", 9:"🥰", 10:"😍"}
radio_buttons = []

for row in range(2):
    frame_row = tk.Frame(emociones_frame, bg=bg_color)
    frame_row.pack()
    for col in range(5):
        valor = row * 5 + col + 1
        rb = tk.Radiobutton(frame_row, text=f"{emociones[valor]} {valor}", variable=var_estado, value=valor,
                            bg=radio_bg, fg=label_color, selectcolor=radio_bg, activebackground=radio_bg,
                            activeforeground=radio_selected, font=emoji_font, indicatoron=0, width=4, relief=tk.RAISED,
                            bd=1, highlightbackground=radio_selected, highlightcolor=radio_selected,
                            highlightthickness=1, overrelief=tk.SOLID)
        rb.pack(side=tk.LEFT, padx=2, pady=2)
        radio_buttons.append(rb)

entry_habitos = crear_campo(main_frame, "5. Hábitos realizados hoy (escribe libremente):", height=2)

btn_guardar = tk.Button(main_frame, text="💾 GUARDAR DIARIO", command=guardar_respuesta, bg=button_bg,
                        fg=button_fg, activebackground=button_active, activeforeground=button_fg,
                        font=button_font, relief=tk.FLAT, padx=20, pady=8, bd=0)
btn_guardar.pack(pady=20)

widgets_tab = [entry_aprendizaje, entry_logro, entry_mejorar, *radio_buttons, entry_habitos, btn_guardar]
for widget in widgets_tab:
    widget.bind("<Tab>", on_tab)
    if isinstance(widget, tk.Text):
        widget.bind("<Shift-Tab>", on_tab)

inicializar_excel()
entry_aprendizaje.focus_set()
centrar_ventana(root)
root.mainloop()
